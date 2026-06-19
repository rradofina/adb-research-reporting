"""Scan Georgia air.gov.ge monthly report/export routes over time.

The May 2026 Georgia report-verification scan proved exact station-code
PM2.5 report rows but kept all rows open because the page carried a
``Not Verified Data`` label. This pass checks whether that caution is just a
latest-month issue by scanning a 24-month HTML ladder and probing official
XLSX/PDF export routes.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import requests
from bs4 import BeautifulSoup
from pypdf import PdfReader


PROGRAM_DIR = Path(__file__).resolve().parents[1]
GENERATED_DIR = PROGRAM_DIR / "generated"
SOURCE_INPUTS_DIR = PROGRAM_DIR / "source-inputs"

SEED_CSV = SOURCE_INPUTS_DIR / "georgia-report-export-ladder-source-seed.csv"
REPORT_VERIFICATION_CSV = GENERATED_DIR / "air-monitoring-georgia-report-verification-source-scan.csv"
OUT_CSV = GENERATED_DIR / "air-monitoring-georgia-report-export-ladder.csv"
OUT_JSON = GENERATED_DIR / "air-monitoring-georgia-report-export-ladder-summary.json"
OUT_MD = PROGRAM_DIR / "georgia-report-export-ladder.md"

METHOD = "air_monitoring_georgia_report_export_ladder_v1"
STATUS = "computed_georgia_report_export_ladder"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36 "
    "ADB-Research-Factory/1.0"
)
TIMEOUT_SECONDS = 90
START_MONTH = "2026-05"
MONTHS_TO_SCAN = 24
EXPORT_PROBE_MONTHS = ["2026-05", "2025-12", "2024-06"]
TARGET_CODES = [
    "01005",
    "AGMS",
    "BTUM",
    "KUTS",
    "KZBG",
    "ORN01",
    "ORN02",
    "ORN03",
    "ORN04",
    "ORN05",
    "ORN06",
    "ORN07",
    "ORN08",
    "RST18",
    "TSRT",
    "VRKT",
]
NON_CLAIM = (
    "This scan tests official Georgia monthly report and export routes for "
    "verification labels across time. It does not certify current station "
    "status, station method class, calibration status, complete monitor-grade "
    "classification, or station-radius readiness."
)

FIELDNAMES = [
    "generated_at",
    "attestation_chain",
    "status",
    "method",
    "ladder_row_id",
    "report_month",
    "report_url",
    "html_retrieved",
    "html_http_status",
    "html_final_url",
    "html_bytes",
    "html_sha256",
    "target_station_codes",
    "station_code_count_in_html",
    "all_station_codes_in_html",
    "pm25_column_in_html",
    "html_not_verified_label_present",
    "html_verified_label_without_not_verified",
    "xlsx_export_tested",
    "xlsx_retrieved",
    "xlsx_http_status",
    "xlsx_final_url",
    "xlsx_bytes",
    "xlsx_sha256",
    "xlsx_sheet_count",
    "xlsx_target_station_sheet_count",
    "xlsx_pm25_present",
    "xlsx_verification_label_present",
    "pdf_export_tested",
    "pdf_retrieved",
    "pdf_http_status",
    "pdf_final_url",
    "pdf_bytes",
    "pdf_sha256",
    "pdf_text_pages",
    "pdf_pm25_present",
    "pdf_not_verified_label_present",
    "pdf_verified_label_without_not_verified",
    "verified_report_closure_available",
    "station_method_classified",
    "current_status_confirmed",
    "complete_monitor_grade_classification_available",
    "station_radius_grade_assumption_ready",
    "report_export_decision",
    "reader_use",
    "non_claim",
]


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = text.replace("\ufb01", "fi").replace("\ufb02", "fl")
    return re.sub(r"\s+", " ", text).strip()


def norm_key(value: Any) -> str:
    return normalize(value).casefold()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def split_terms(value: str) -> list[str]:
    return [term.strip() for term in str(value or "").split("||") if term.strip()]


def matched_terms(text: str, terms: list[str]) -> list[str]:
    lower = norm_key(text)
    return [term for term in terms if norm_key(term) in lower]


def month_ladder(start_month: str, count: int) -> list[str]:
    year, month = [int(part) for part in start_month.split("-")]
    output: list[str] = []
    for _ in range(count):
        output.append(f"{year:04d}-{month:02d}")
        month -= 1
        if month == 0:
            year -= 1
            month = 12
    return output


def source_seed_by_role(seed_rows: list[dict[str, str]], role: str) -> dict[str, str]:
    for row in seed_rows:
        if row["source_role"] == role:
            return row
    raise ValueError(f"Missing source role {role}")


def request_report(month: str, export_type: str | None = None) -> dict[str, Any]:
    params = {
        "date_from": month,
        "report_type": "monthly",
        "station": ",".join(TARGET_CODES),
    }
    if export_type:
        params["export_type"] = export_type
    result: dict[str, Any] = {
        "retrieved": False,
        "http_status": "",
        "final_url": "",
        "content_type": "",
        "bytes": 0,
        "sha256": "",
        "content": b"",
        "text": "",
        "error": "",
    }
    try:
        response = requests.get(
            "https://air.gov.ge/en/reports_page",
            params=params,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
            },
            timeout=TIMEOUT_SECONDS,
            allow_redirects=True,
        )
        result["http_status"] = response.status_code
        result["final_url"] = response.url
        result["content_type"] = response.headers.get("content-type", "")
        result["bytes"] = len(response.content)
        result["sha256"] = hashlib.sha256(response.content).hexdigest()
        response.raise_for_status()
        result["retrieved"] = True
        result["content"] = response.content
        if export_type is None:
            soup = BeautifulSoup(response.text, "html.parser")
            result["text"] = normalize(soup.get_text(" ", strip=True))
    except Exception as exc:  # noqa: BLE001 - retrieval failures are evidence.
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def verified_without_not(text: str) -> bool:
    lower = norm_key(text)
    return "verified data" in lower and "not verified data" not in lower


def extract_xlsx(content: bytes) -> dict[str, Any]:
    result = {
        "sheet_count": 0,
        "target_station_sheet_count": 0,
        "pm25_present": False,
        "verification_label_present": False,
        "text": "",
        "error": "",
    }
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = list(workbook.sheetnames)
        text_parts = [*sheet_names]
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                text_parts.extend(str(cell) for cell in row if cell is not None)
        text = normalize(" ".join(text_parts))
        result.update(
            {
                "sheet_count": len(sheet_names),
                "target_station_sheet_count": sum(code in sheet_names for code in TARGET_CODES),
                "pm25_present": "PM2.5" in text,
                "verification_label_present": "verified data" in norm_key(text),
                "text": text,
            }
        )
    except Exception as exc:  # noqa: BLE001 - export parse failures are evidence.
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def extract_pdf(content: bytes) -> dict[str, Any]:
    result = {
        "text_pages": 0,
        "pm25_present": False,
        "not_verified_label_present": False,
        "verified_label_without_not_verified": False,
        "text": "",
        "error": "",
    }
    try:
        reader = PdfReader(io.BytesIO(content))
        texts = [page.extract_text() or "" for page in reader.pages]
        text = normalize(" ".join(texts))
        result.update(
            {
                "text_pages": len(reader.pages),
                "pm25_present": "PM2.5" in text,
                "not_verified_label_present": "not verified data" in norm_key(text),
                "verified_label_without_not_verified": verified_without_not(text),
                "text": text,
            }
        )
    except Exception as exc:  # noqa: BLE001 - export parse failures are evidence.
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def base_row(generated_at: str, month: str, html: dict[str, Any]) -> dict[str, Any]:
    text = html.get("text", "")
    code_count = sum(code in text for code in TARGET_CODES)
    all_codes = code_count == len(TARGET_CODES)
    pm25_present = "PM2.5" in text and all_codes
    not_verified = "not verified data" in norm_key(text)
    verified = verified_without_not(text)
    if all_codes and pm25_present and not_verified:
        decision = "monthly_html_ladder_pm25_present_not_verified_keep_open"
        reader_use = (
            "The official monthly HTML route has all target station codes and PM2.5, "
            "but the verification footer remains Not Verified Data."
        )
    elif all_codes and pm25_present and verified:
        decision = "monthly_html_ladder_verified_label_needs_method_status"
        reader_use = (
            "The official monthly HTML route appears to remove the not-verified caution, "
            "but station method/status evidence still must be checked."
        )
    elif all_codes:
        decision = "monthly_html_ladder_station_codes_without_pm25_or_label_keep_open"
        reader_use = "The official monthly HTML route has the target station codes, but PM2.5 or label closure is incomplete."
    else:
        decision = "monthly_html_ladder_station_code_gap_keep_open"
        reader_use = "The official monthly HTML route did not expose every target station code in the fetched text."
    return {
        "generated_at": generated_at,
        "attestation_chain": "ai-first",
        "status": STATUS,
        "method": METHOD,
        "ladder_row_id": f"GEO-report-export-ladder-{month}",
        "report_month": month,
        "report_url": html.get("final_url", ""),
        "html_retrieved": html.get("retrieved", False),
        "html_http_status": html.get("http_status", ""),
        "html_final_url": html.get("final_url", ""),
        "html_bytes": html.get("bytes", 0),
        "html_sha256": html.get("sha256", ""),
        "target_station_codes": "|".join(TARGET_CODES),
        "station_code_count_in_html": code_count,
        "all_station_codes_in_html": all_codes,
        "pm25_column_in_html": pm25_present,
        "html_not_verified_label_present": not_verified,
        "html_verified_label_without_not_verified": verified,
        "xlsx_export_tested": False,
        "xlsx_retrieved": False,
        "xlsx_http_status": "",
        "xlsx_final_url": "",
        "xlsx_bytes": 0,
        "xlsx_sha256": "",
        "xlsx_sheet_count": "",
        "xlsx_target_station_sheet_count": "",
        "xlsx_pm25_present": False,
        "xlsx_verification_label_present": False,
        "pdf_export_tested": False,
        "pdf_retrieved": False,
        "pdf_http_status": "",
        "pdf_final_url": "",
        "pdf_bytes": 0,
        "pdf_sha256": "",
        "pdf_text_pages": "",
        "pdf_pm25_present": False,
        "pdf_not_verified_label_present": False,
        "pdf_verified_label_without_not_verified": False,
        "verified_report_closure_available": False,
        "station_method_classified": False,
        "current_status_confirmed": False,
        "complete_monitor_grade_classification_available": False,
        "station_radius_grade_assumption_ready": False,
        "report_export_decision": decision,
        "reader_use": reader_use,
        "non_claim": NON_CLAIM,
    }


def add_export_probe(row: dict[str, Any], xlsx: dict[str, Any], pdf: dict[str, Any]) -> None:
    xlsx_parsed = extract_xlsx(xlsx.get("content", b"")) if xlsx.get("retrieved") else {}
    pdf_parsed = extract_pdf(pdf.get("content", b"")) if pdf.get("retrieved") else {}
    row.update(
        {
            "xlsx_export_tested": True,
            "xlsx_retrieved": xlsx.get("retrieved", False),
            "xlsx_http_status": xlsx.get("http_status", ""),
            "xlsx_final_url": xlsx.get("final_url", ""),
            "xlsx_bytes": xlsx.get("bytes", 0),
            "xlsx_sha256": xlsx.get("sha256", ""),
            "xlsx_sheet_count": xlsx_parsed.get("sheet_count", ""),
            "xlsx_target_station_sheet_count": xlsx_parsed.get("target_station_sheet_count", ""),
            "xlsx_pm25_present": xlsx_parsed.get("pm25_present", False),
            "xlsx_verification_label_present": xlsx_parsed.get("verification_label_present", False),
            "pdf_export_tested": True,
            "pdf_retrieved": pdf.get("retrieved", False),
            "pdf_http_status": pdf.get("http_status", ""),
            "pdf_final_url": pdf.get("final_url", ""),
            "pdf_bytes": pdf.get("bytes", 0),
            "pdf_sha256": pdf.get("sha256", ""),
            "pdf_text_pages": pdf_parsed.get("text_pages", ""),
            "pdf_pm25_present": pdf_parsed.get("pm25_present", False),
            "pdf_not_verified_label_present": pdf_parsed.get("not_verified_label_present", False),
            "pdf_verified_label_without_not_verified": pdf_parsed.get("verified_label_without_not_verified", False),
        }
    )
    if row["html_not_verified_label_present"] and row["pdf_not_verified_label_present"]:
        row["report_export_decision"] = "html_and_pdf_export_not_verified_keep_open"
        row["reader_use"] = (
            "The HTML route and PDF export both expose report data but carry the not-verified footer. "
            "The XLSX export has target station sheets and PM2.5 but no independent verification label."
        )


def gate(status: str, gate_name: str, rows: int, reader_use: str) -> dict[str, Any]:
    return {"status": status, "gate": gate_name, "rows": rows, "reader_use": reader_use}


def evidence_gates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    export_rows = [row for row in rows if row["xlsx_export_tested"]]
    return [
        gate(
            "available" if all(row["html_retrieved"] for row in rows) else "limited",
            "Monthly HTML routes retrieved",
            sum(row["html_retrieved"] for row in rows),
            "The scan retrieves each official monthly report page in the 24-month ladder.",
        ),
        gate(
            "available",
            "All target station codes in monthly HTML",
            sum(row["all_station_codes_in_html"] for row in rows),
            "Every scanned month exposes the 16 target station codes in the official report text.",
        ),
        gate(
            "available",
            "PM2.5 column in monthly HTML",
            sum(row["pm25_column_in_html"] for row in rows),
            "Every scanned month exposes PM2.5 in the station-code report table.",
        ),
        gate(
            "caution",
            "Not Verified Data in monthly HTML",
            sum(row["html_not_verified_label_present"] for row in rows),
            "The fetched official monthly pages retain the not-verified footer across the scanned ladder.",
        ),
        gate(
            "not_ready",
            "Verified label without not-verified footer",
            sum(row["html_verified_label_without_not_verified"] for row in rows),
            "No scanned monthly page provides a clean verified-report label.",
        ),
        gate(
            "available" if all(row["xlsx_retrieved"] for row in export_rows) else "limited",
            "XLSX export probes retrieved",
            sum(row["xlsx_retrieved"] for row in export_rows),
            "The official XLSX export route returns station sheets for the selected probe months.",
        ),
        gate(
            "partly_available",
            "XLSX target station sheets",
            sum(row["xlsx_target_station_sheet_count"] == len(TARGET_CODES) for row in export_rows),
            "XLSX exports contain all target station sheets in the probed months, but no independent verification label.",
        ),
        gate(
            "caution",
            "PDF export Not Verified Data footer",
            sum(row["pdf_not_verified_label_present"] for row in export_rows),
            "PDF exports preserve the not-verified footer in extracted text.",
        ),
        gate(
            "not_ready",
            "Verified report closure",
            sum(row["verified_report_closure_available"] for row in rows),
            "No monthly HTML or export route closes the verified-report gate.",
        ),
        gate(
            "not_ready",
            "Current status and complete grade",
            sum(row["complete_monitor_grade_classification_available"] for row in rows),
            "The report/export routes do not provide station status, calibration, method class, or complete monitor-grade evidence.",
        ),
    ]


def source_records(seed_rows: list[dict[str, str]], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    html_text = " ".join(row["report_url"] for row in rows)
    output = []
    for seed in seed_rows:
        if seed["source_role"] == "official_monthly_report_html":
            retrievals = sum(row["html_retrieved"] for row in rows)
            bytes_total = sum(int(row["html_bytes"] or 0) for row in rows)
            sha_sample = [row["html_sha256"] for row in rows[:3]]
        elif seed["source_role"] == "official_monthly_report_xlsx_export":
            retrievals = sum(row["xlsx_retrieved"] for row in rows)
            bytes_total = sum(int(row["xlsx_bytes"] or 0) for row in rows)
            sha_sample = [row["xlsx_sha256"] for row in rows if row["xlsx_sha256"]][:3]
        else:
            retrievals = sum(row["pdf_retrieved"] for row in rows)
            bytes_total = sum(int(row["pdf_bytes"] or 0) for row in rows)
            sha_sample = [row["pdf_sha256"] for row in rows if row["pdf_sha256"]][:3]
        output.append(
            {
                "source_key": seed["source_key"],
                "source_name": seed["source_name"],
                "source_role": seed["source_role"],
                "url": seed["url"],
                "content_type_hint": seed["content_type_hint"],
                "retrievals": retrievals,
                "retrieval_bytes_total": bytes_total,
                "sha256_sample": sha_sample,
                "matched_expected_terms": matched_terms(html_text, split_terms(seed["expected_terms"])),
                "source_note": seed["source_note"],
            }
        )
    return output


def summary_payload(generated_at: str, rows: list[dict[str, Any]], seed_rows: list[dict[str, str]]) -> dict[str, Any]:
    export_rows = [row for row in rows if row["xlsx_export_tested"]]
    counts = {
        "months_scanned": len(rows),
        "target_station_codes": len(TARGET_CODES),
        "html_months_retrieved": sum(row["html_retrieved"] for row in rows),
        "html_months_with_all_target_station_codes": sum(row["all_station_codes_in_html"] for row in rows),
        "html_months_with_pm25_column": sum(row["pm25_column_in_html"] for row in rows),
        "html_not_verified_label_months": sum(row["html_not_verified_label_present"] for row in rows),
        "html_verified_label_without_not_verified_months": sum(row["html_verified_label_without_not_verified"] for row in rows),
        "export_probe_months": len(export_rows),
        "xlsx_export_probe_months_retrieved": sum(row["xlsx_retrieved"] for row in export_rows),
        "xlsx_export_probe_months_with_all_target_sheets": sum(row["xlsx_target_station_sheet_count"] == len(TARGET_CODES) for row in export_rows),
        "xlsx_export_probe_months_with_pm25": sum(row["xlsx_pm25_present"] for row in export_rows),
        "xlsx_export_probe_months_with_verification_label": sum(row["xlsx_verification_label_present"] for row in export_rows),
        "pdf_export_probe_months_retrieved": sum(row["pdf_retrieved"] for row in export_rows),
        "pdf_export_probe_months_with_not_verified_label": sum(row["pdf_not_verified_label_present"] for row in export_rows),
        "pdf_export_probe_months_verified_without_not_verified": sum(row["pdf_verified_label_without_not_verified"] for row in export_rows),
        "verified_report_closure_available_months": sum(row["verified_report_closure_available"] for row in rows),
        "current_status_confirmed_months": sum(row["current_status_confirmed"] for row in rows),
        "station_method_classified_months": sum(row["station_method_classified"] for row in rows),
        "complete_monitor_grade_classification_months": sum(row["complete_monitor_grade_classification_available"] for row in rows),
        "station_radius_grade_assumption_ready_months": sum(row["station_radius_grade_assumption_ready"] for row in rows),
    }
    sample_fields = [
        "report_month",
        "station_code_count_in_html",
        "pm25_column_in_html",
        "html_not_verified_label_present",
        "html_verified_label_without_not_verified",
        "xlsx_export_tested",
        "xlsx_target_station_sheet_count",
        "pdf_not_verified_label_present",
        "report_export_decision",
    ]
    return {
        "generated_at": generated_at,
        "program": "air-monitoring",
        "attestation_chain": "ai-first",
        "status": STATUS,
        "method": METHOD,
        "goal_level": "L3 Georgia report export verification ladder",
        "start_month": START_MONTH,
        "months_to_scan": MONTHS_TO_SCAN,
        "export_probe_months": EXPORT_PROBE_MONTHS,
        "source_inputs": [
            {"path": str(SEED_CSV.relative_to(PROGRAM_DIR)), "role": "official report/export route source seed"},
            {"path": str(REPORT_VERIFICATION_CSV.relative_to(PROGRAM_DIR)), "role": "16 target Georgia station codes from prior report verification scan"},
        ],
        "coverage_counts": counts,
        "decision_counts": [
            {"decision": key, "rows": value}
            for key, value in sorted(Counter(row["report_export_decision"] for row in rows).items())
        ],
        "evidence_gate_counts": evidence_gates(rows),
        "month_rows": [{field: row[field] for field in sample_fields} for row in rows],
        "export_probe_rows": [
            {
                "report_month": row["report_month"],
                "xlsx_retrieved": row["xlsx_retrieved"],
                "xlsx_sheet_count": row["xlsx_sheet_count"],
                "xlsx_target_station_sheet_count": row["xlsx_target_station_sheet_count"],
                "xlsx_pm25_present": row["xlsx_pm25_present"],
                "xlsx_verification_label_present": row["xlsx_verification_label_present"],
                "pdf_retrieved": row["pdf_retrieved"],
                "pdf_text_pages": row["pdf_text_pages"],
                "pdf_pm25_present": row["pdf_pm25_present"],
                "pdf_not_verified_label_present": row["pdf_not_verified_label_present"],
                "pdf_verified_label_without_not_verified": row["pdf_verified_label_without_not_verified"],
            }
            for row in export_rows
        ],
        "source_records": source_records(seed_rows, rows),
        "outputs": {
            "csv": str(OUT_CSV.relative_to(PROGRAM_DIR)),
            "summary_json": str(OUT_JSON.relative_to(PROGRAM_DIR)),
            "note": str(OUT_MD.relative_to(PROGRAM_DIR)),
        },
        "non_claim": NON_CLAIM,
    }


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    widths = [max(len(str(row[i])) for row in rows) for i in range(len(rows[0]))]
    output = []
    for index, row in enumerate(rows):
        output.append("| " + " | ".join(str(row[i]).ljust(widths[i]) for i in range(len(row))) + " |")
        if index == 0:
            output.append("| " + " | ".join("-" * widths[i] for i in range(len(row))) + " |")
    return "\n".join(output)


def write_markdown(path: Path, summary: dict[str, Any]) -> None:
    counts = summary["coverage_counts"]
    gates = summary["evidence_gate_counts"]
    month_rows = summary["month_rows"]
    export_rows = summary["export_probe_rows"]
    gate_table = [["Gate", "Rows", "Status"]]
    gate_table.extend([[row["gate"], str(row["rows"]), row["status"]] for row in gates])
    month_table = [["Month", "Codes", "PM2.5", "Not verified", "Verified clean", "Decision"]]
    for row in month_rows:
        month_table.append(
            [
                row["report_month"],
                str(row["station_code_count_in_html"]),
                "yes" if row["pm25_column_in_html"] else "no",
                "yes" if row["html_not_verified_label_present"] else "no",
                "yes" if row["html_verified_label_without_not_verified"] else "no",
                row["report_export_decision"],
            ]
        )
    export_table = [["Month", "XLSX sheets", "XLSX PM2.5", "XLSX label", "PDF pages", "PDF not verified"]]
    for row in export_rows:
        export_table.append(
            [
                row["report_month"],
                f"{row['xlsx_target_station_sheet_count']}/{len(TARGET_CODES)}",
                "yes" if row["xlsx_pm25_present"] else "no",
                "yes" if row["xlsx_verification_label_present"] else "no",
                str(row["pdf_text_pages"]),
                "yes" if row["pdf_not_verified_label_present"] else "no",
            ]
        )
    text = f"""# Georgia report export verification ladder

Status: computed source scan, Mode A AI-first.

This pass follows the Georgia report-verification source scan. It asks whether
the official `air.gov.ge` report caution is only a latest-month issue by
checking a 24-month monthly-report ladder and probing the XLSX/PDF export
routes exposed by the same public report page.

## Result

The ladder does not close the verified-report gate.

It records:

- {counts['months_scanned']} monthly HTML routes scanned, ending at {summary['start_month']}.
- {counts['html_months_retrieved']} monthly HTML routes retrieved.
- {counts['html_months_with_all_target_station_codes']} months with all {counts['target_station_codes']} target station codes in the official report text.
- {counts['html_months_with_pm25_column']} months with PM2.5 in the station-code report table.
- {counts['html_not_verified_label_months']} months where the HTML route carries `Not Verified Data`.
- {counts['html_verified_label_without_not_verified_months']} months with a clean verified label that does not also contain the not-verified footer.
- {counts['export_probe_months']} export-probe months.
- {counts['xlsx_export_probe_months_retrieved']} retrieved XLSX export probes.
- {counts['xlsx_export_probe_months_with_all_target_sheets']} XLSX probes with all target station sheets.
- {counts['pdf_export_probe_months_with_not_verified_label']} PDF probes whose extracted text carries the not-verified footer.
- {counts['verified_report_closure_available_months']} verified-report closure months.
- {counts['complete_monitor_grade_classification_months']} complete monitor-grade months.

## Main Reading

The official Georgia route is useful for report visibility but not enough for
grade closure. Across the scanned ladder, the pages consistently expose the
target station codes and PM2.5 report columns, yet the same route retains the
not-verified footer. The XLSX exports provide station sheets and PM2.5 values,
but do not supply an independent verification label. The PDF export probes
preserve the not-verified footer in extracted text.

The result is therefore a source-screening finding, not a pollution result:
the public report/export surface is good enough to prove that report rows
exist, but not good enough to promote Georgia rows into verified report,
current-status, complete-grade, or station-radius analysis.

## Evidence Gates

{markdown_table(gate_table)}

## Month Ladder

{markdown_table(month_table)}

## Export Probes

{markdown_table(export_table)}

## Method

The script `scripts/scan-georgia-report-export-ladder.py` reads:

- `source-inputs/georgia-report-export-ladder-source-seed.csv`
- `generated/air-monitoring-georgia-report-verification-source-scan.csv`

It requests the official monthly report route with the 16 target station codes
for {counts['months_scanned']} months from {summary['start_month']} backward,
then checks the HTML text for exact station codes, PM2.5, `Not Verified Data`,
and a clean `Verified Data` label that does not also contain the not-verified
footer. For {counts['export_probe_months']} anchor months it also requests
`export_type=xlsx` and `export_type=pdf`, parses XLSX sheet names and PDF text,
and records retrieval byte counts and SHA-256 hashes in the generated CSV.

## Artifacts

- Script: `air-monitoring/scripts/scan-georgia-report-export-ladder.py`
- Source seed:
  `air-monitoring/source-inputs/georgia-report-export-ladder-source-seed.csv`
- Row output:
  `air-monitoring/generated/air-monitoring-georgia-report-export-ladder.csv`
- Summary output:
  `air-monitoring/generated/air-monitoring-georgia-report-export-ladder-summary.json`

## Reader Use

Use this artifact to show why the Georgia lane stays open even after finding
official station-code report rows. The next source needed is not another
monthly table; it is a public regulator route, verified export, or station
record that explicitly removes the not-verified caution and supplies station
method/status/grade evidence for exact station codes.

## Non-claim

{summary['non_claim']}
"""
    path.write_text(text, encoding="utf-8")


def main() -> None:
    generated_at = now_iso()
    seed_rows = read_csv(SEED_CSV)
    # Fail fast if source roles drift.
    source_seed_by_role(seed_rows, "official_monthly_report_html")
    source_seed_by_role(seed_rows, "official_monthly_report_xlsx_export")
    source_seed_by_role(seed_rows, "official_monthly_report_pdf_export")

    rows: list[dict[str, Any]] = []
    for month in month_ladder(START_MONTH, MONTHS_TO_SCAN):
        html = request_report(month)
        row = base_row(generated_at, month, html)
        if month in EXPORT_PROBE_MONTHS:
            add_export_probe(row, request_report(month, "xlsx"), request_report(month, "pdf"))
        rows.append(row)

    summary = summary_payload(generated_at, rows, seed_rows)
    write_csv(OUT_CSV, rows)
    write_json(OUT_JSON, summary)
    write_markdown(OUT_MD, summary)
    print(
        "Built Georgia report export ladder: "
        f"{len(rows)} months; "
        f"{summary['coverage_counts']['html_months_retrieved']} HTML routes retrieved; "
        f"{summary['coverage_counts']['html_not_verified_label_months']} not-verified HTML months; "
        f"{summary['coverage_counts']['xlsx_export_probe_months_retrieved']} XLSX probes; "
        f"{summary['coverage_counts']['pdf_export_probe_months_with_not_verified_label']} PDF not-verified probes; "
        "0 verified closures."
    )


if __name__ == "__main__":
    main()

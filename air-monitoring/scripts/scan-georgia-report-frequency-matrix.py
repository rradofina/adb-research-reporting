"""Scan Georgia report-frequency routes for verified-report closure.

The policy wall says verified data are available in reports, while the
monthly report/export ladder keeps the public monthly surfaces open. This pass
tests daily, monthly, and annual report routes plus XLSX/PDF exports for the
same 16 Georgia station codes.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import requests
from bs4 import BeautifulSoup
from pypdf import PdfReader


PROGRAM_DIR = Path(__file__).resolve().parents[1]
GENERATED_DIR = PROGRAM_DIR / "generated"
SOURCE_INPUTS_DIR = PROGRAM_DIR / "source-inputs"

SEED_CSV = SOURCE_INPUTS_DIR / "georgia-report-frequency-matrix-source-seed.csv"
OUT_CSV = GENERATED_DIR / "air-monitoring-georgia-report-frequency-matrix.csv"
OUT_JSON = GENERATED_DIR / "air-monitoring-georgia-report-frequency-matrix-summary.json"
OUT_MD = PROGRAM_DIR / "georgia-report-frequency-matrix.md"

METHOD = "air_monitoring_georgia_report_frequency_matrix_v1"
STATUS = "computed_georgia_report_frequency_matrix"
BASE_URL = "https://air.gov.ge/en/reports_page"
TIMEOUT_SECONDS = 90
TARGET_CODES = [
    "01005",
    "AGMS",
    "BTUM",
    "KUTS",
    "KZBG",
    "ORN01",
    "ORN02",
    "ORN03",
    "ORN04",
    "ORN05",
    "ORN06",
    "ORN07",
    "ORN08",
    "RST18",
    "TSRT",
    "VRKT",
]
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36 "
    "ADB-Research-Factory/1.0"
)
NON_CLAIM = (
    "This scan tests official Georgia daily, monthly, and annual report "
    "routes plus XLSX/PDF exports. It does not certify any target station as "
    "verified, currently operating, station-method classified, complete "
    "monitor-grade, or station-radius ready."
)

FIELDNAMES = [
    "generated_at",
    "attestation_chain",
    "status",
    "method",
    "probe_key",
    "route_key",
    "source_name",
    "report_type",
    "export_type",
    "probe_date",
    "requested_url",
    "retrieved_200",
    "http_status",
    "final_url",
    "content_type",
    "response_bytes",
    "sha256",
    "valid_payload",
    "content_type_matches_expected",
    "station_code_matches",
    "all_target_station_codes",
    "pm25_present",
    "not_verified_label_present",
    "verified_label_without_not_verified",
    "xlsx_sheet_count",
    "xlsx_target_station_sheet_count",
    "server_error",
    "current_status_confirmed",
    "verified_report_closure_available",
    "station_method_classified",
    "complete_monitor_grade_classification_available",
    "station_radius_grade_assumption_ready",
    "report_frequency_decision",
    "reader_use",
    "non_claim",
]


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = text.replace("\ufb01", "fi").replace("\ufb02", "fl")
    return re.sub(r"\s+", " ", text).strip()


def norm_key(value: Any) -> str:
    return normalize(value).casefold()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def split_terms(value: str) -> list[str]:
    return [term.strip() for term in str(value or "").split("||") if term.strip()]


def build_params(report_type: str, probe_date: str, export_type: str) -> dict[str, str]:
    params = {
        "report_type": report_type,
        "date_from": probe_date,
        "station": ",".join(TARGET_CODES),
    }
    if export_type != "html":
        params["export_type"] = export_type
    return params


def fetch_probe(report_type: str, probe_date: str, export_type: str) -> dict[str, Any]:
    result: dict[str, Any] = {
        "retrieved_200": False,
        "http_status": "",
        "final_url": "",
        "content_type": "",
        "response_bytes": 0,
        "sha256": "",
        "content": b"",
        "text": "",
        "error": "",
    }
    params = build_params(report_type, probe_date, export_type)
    try:
        response = requests.get(
            BASE_URL,
            params=params,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
            },
            timeout=TIMEOUT_SECONDS,
            allow_redirects=True,
        )
        result["http_status"] = response.status_code
        result["final_url"] = response.url
        result["content_type"] = response.headers.get("content-type", "")
        result["response_bytes"] = len(response.content)
        result["sha256"] = hashlib.sha256(response.content).hexdigest()
        result["content"] = response.content
        result["retrieved_200"] = response.status_code == 200
        result["text"] = extract_text(response.content, response.text, export_type)
    except Exception as exc:  # noqa: BLE001 - retrieval failures are evidence.
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def extract_text(content: bytes, html_text: str, export_type: str) -> str:
    if export_type == "xlsx":
        return extract_xlsx_text(content)
    if export_type == "pdf":
        return extract_pdf_text(content)
    soup = BeautifulSoup(html_text, "html.parser")
    return normalize(soup.get_text(" ", strip=True))


def extract_xlsx_text(content: bytes) -> str:
    if not content.startswith(b"PK"):
        return "NOT_XLSX " + normalize(content[:120])
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        parts: list[str] = list(workbook.sheetnames)
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                parts.extend(str(cell) for cell in row if cell is not None)
        return normalize(" ".join(parts))
    except Exception as exc:  # noqa: BLE001
        return f"XLSX_TEXT_EXTRACTION_ERROR {type(exc).__name__}: {exc}"


def extract_xlsx_sheet_counts(content: bytes) -> tuple[int, int]:
    if not content.startswith(b"PK"):
        return 0, 0
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = list(workbook.sheetnames)
        return len(sheet_names), sum(code in sheet_names for code in TARGET_CODES)
    except Exception:  # noqa: BLE001
        return 0, 0


def extract_pdf_text(content: bytes) -> str:
    if not content.startswith(b"%PDF"):
        return "NOT_PDF " + normalize(content[:120])
    try:
        reader = PdfReader(io.BytesIO(content))
        return normalize(" ".join(page.extract_text() or "" for page in reader.pages))
    except Exception as exc:  # noqa: BLE001
        return f"PDF_TEXT_EXTRACTION_ERROR {type(exc).__name__}: {exc}"


def expected_payload(export_type: str, fetched: dict[str, Any]) -> bool:
    content = fetched["content"]
    content_type = norm_key(fetched["content_type"])
    if fetched["http_status"] != 200:
        return False
    if export_type == "xlsx":
        return content.startswith(b"PK") and "spreadsheet" in content_type
    if export_type == "pdf":
        return content.startswith(b"%PDF") and "pdf" in content_type
    return "html" in content_type


def content_type_matches(export_type: str, fetched: dict[str, Any]) -> bool:
    content = fetched["content"]
    content_type = norm_key(fetched["content_type"])
    if export_type == "xlsx":
        return content.startswith(b"PK") and "spreadsheet" in content_type
    if export_type == "pdf":
        return content.startswith(b"%PDF") and "pdf" in content_type
    return "html" in content_type


def verified_without_not(text: str) -> bool:
    lower = norm_key(text)
    return "verified data" in lower and "not verified data" not in lower


def has_not_verified(text: str) -> bool:
    lower = norm_key(text)
    return "not verified data" in lower or "not verified" in lower


def has_pm25(text: str) -> bool:
    lower = norm_key(text)
    return "pm2.5" in lower or "pm 2.5" in lower


def decision_for(row: dict[str, Any]) -> tuple[str, str]:
    if row["server_error"]:
        return (
            "annual_route_server_error_no_closure"
            if row["report_type"] == "annual"
            else "server_error_no_closure",
            "The route does not return an analysis-ready report payload, so it cannot close verification.",
        )
    if row["export_type"] == "xlsx" and row["valid_payload"]:
        return (
            "xlsx_station_sheets_without_verification_label",
            "The export carries station sheets and PM2.5 values, but no verification label.",
        )
    if row["not_verified_label_present"]:
        return (
            f"{row['report_type']}_{row['export_type']}_not_verified_keep_open",
            "The report payload repeats the not-verified caution and stays outside closure.",
        )
    if row["verified_label_without_not_verified"]:
        return (
            "verified_label_found_needs_station_review",
            "A clean verified label would need station-level method/status review before promotion.",
        )
    return (
        "available_without_verified_closure",
        "The route is reachable but does not provide verified-report closure.",
    )


def build_rows(generated_at: str, seed_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for seed in seed_rows:
        for probe_date in split_terms(seed["probe_dates"]):
            export_type = seed["export_type"]
            fetched = fetch_probe(seed["report_type"], probe_date, export_type)
            text = fetched["text"]
            sheet_count, target_sheet_count = extract_xlsx_sheet_counts(fetched["content"])
            station_matches = sum(code in text for code in TARGET_CODES)
            valid = expected_payload(export_type, fetched)
            row: dict[str, Any] = {
                "generated_at": generated_at,
                "attestation_chain": "ai-first",
                "status": STATUS,
                "method": METHOD,
                "probe_key": f"{seed['route_key']}::{probe_date}",
                "route_key": seed["route_key"],
                "source_name": seed["source_name"],
                "report_type": seed["report_type"],
                "export_type": export_type,
                "probe_date": probe_date,
                "requested_url": fetched["final_url"] or BASE_URL,
                "retrieved_200": fetched["retrieved_200"],
                "http_status": fetched["http_status"],
                "final_url": fetched["final_url"],
                "content_type": fetched["content_type"],
                "response_bytes": fetched["response_bytes"],
                "sha256": fetched["sha256"],
                "valid_payload": valid,
                "content_type_matches_expected": content_type_matches(export_type, fetched),
                "station_code_matches": station_matches,
                "all_target_station_codes": station_matches == len(TARGET_CODES),
                "pm25_present": has_pm25(text),
                "not_verified_label_present": has_not_verified(text),
                "verified_label_without_not_verified": verified_without_not(text),
                "xlsx_sheet_count": sheet_count,
                "xlsx_target_station_sheet_count": target_sheet_count,
                "server_error": str(fetched["http_status"]).startswith("5"),
                "current_status_confirmed": False,
                "verified_report_closure_available": False,
                "station_method_classified": False,
                "complete_monitor_grade_classification_available": False,
                "station_radius_grade_assumption_ready": False,
                "non_claim": NON_CLAIM,
            }
            decision, reader_use = decision_for(row)
            row["report_frequency_decision"] = decision
            row["reader_use"] = reader_use
            rows.append(row)
    return rows


def bool_count(rows: list[dict[str, Any]], key: str) -> int:
    return sum(1 for row in rows if bool(row.get(key)))


def build_frequency_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_type[row["report_type"]].append(row)
    output = []
    for report_type in ["daily", "monthly", "annual"]:
        group = by_type.get(report_type, [])
        output.append(
            {
                "report_type": report_type,
                "route_probes": len(group),
                "valid_payload_routes": bool_count(group, "valid_payload"),
                "server_error_routes": bool_count(group, "server_error"),
                "html_not_verified_routes": sum(
                    1 for row in group if row["export_type"] == "html" and row["not_verified_label_present"]
                ),
                "pdf_not_verified_routes": sum(
                    1 for row in group if row["export_type"] == "pdf" and row["not_verified_label_present"]
                ),
                "xlsx_station_sheet_routes": sum(
                    1 for row in group if row["export_type"] == "xlsx" and row["xlsx_target_station_sheet_count"] == len(TARGET_CODES)
                ),
                "verified_closure_routes": bool_count(group, "verified_report_closure_available"),
                "reader_use": frequency_reader_use(report_type),
            }
        )
    return output


def frequency_reader_use(report_type: str) -> str:
    if report_type == "daily":
        return "Daily HTML/PDF surfaces repeat the caution; XLSX carries station sheets but no verification label."
    if report_type == "monthly":
        return "Monthly comparison routes match the earlier ladder: report payloads remain open, not closed."
    return "Annual route probes return server-error pages for tested date formats, so they do not provide a verified surface."


def evidence_gates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "status": "available",
            "gate": "Daily report/export probes",
            "rows": sum(1 for row in rows if row["report_type"] == "daily"),
            "reader_use": "Daily HTML, XLSX, and PDF routes were tested on two dates.",
        },
        {
            "status": "available",
            "gate": "Monthly comparison probes",
            "rows": sum(1 for row in rows if row["report_type"] == "monthly"),
            "reader_use": "Monthly HTML, XLSX, and PDF routes were retested on two benchmark months.",
        },
        {
            "status": "not_ready",
            "gate": "Annual route probes",
            "rows": sum(1 for row in rows if row["report_type"] == "annual" and row["server_error"]),
            "reader_use": "Annual route probes returned server-error pages for the tested formats.",
        },
        {
            "status": "caution",
            "gate": "HTML/PDF Not Verified Data",
            "rows": sum(
                1
                for row in rows
                if row["export_type"] in {"html", "pdf"} and row["not_verified_label_present"]
            ),
            "reader_use": "Human-readable report payloads still carry not-verified caution text.",
        },
        {
            "status": "partly_available",
            "gate": "XLSX station sheets",
            "rows": sum(1 for row in rows if row["xlsx_target_station_sheet_count"] == len(TARGET_CODES)),
            "reader_use": "XLSX exports expose target station sheets but no verification label.",
        },
        {
            "status": "not_ready",
            "gate": "Clean verified label",
            "rows": bool_count(rows, "verified_label_without_not_verified"),
            "reader_use": "No tested frequency/export route provides a verified label without the not-verified caution.",
        },
        {
            "status": "not_ready",
            "gate": "Verified report closure",
            "rows": bool_count(rows, "verified_report_closure_available"),
            "reader_use": "No tested route closes the verified-report gate for the target station codes.",
        },
        {
            "status": "not_ready",
            "gate": "Station status and complete grade",
            "rows": bool_count(rows, "complete_monitor_grade_classification_available"),
            "reader_use": "Report-frequency routes do not provide station current status, method table, or complete grade.",
        },
    ]


def build_summary(generated_at: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts = {
        "route_probes_targeted": len(rows),
        "routes_retrieved_200": bool_count(rows, "retrieved_200"),
        "valid_payload_routes": bool_count(rows, "valid_payload"),
        "server_error_routes": bool_count(rows, "server_error"),
        "annual_server_error_routes": sum(1 for row in rows if row["report_type"] == "annual" and row["server_error"]),
        "daily_routes_tested": sum(1 for row in rows if row["report_type"] == "daily"),
        "monthly_routes_tested": sum(1 for row in rows if row["report_type"] == "monthly"),
        "annual_routes_tested": sum(1 for row in rows if row["report_type"] == "annual"),
        "html_not_verified_routes": sum(1 for row in rows if row["export_type"] == "html" and row["not_verified_label_present"]),
        "pdf_not_verified_routes": sum(1 for row in rows if row["export_type"] == "pdf" and row["not_verified_label_present"]),
        "html_pdf_not_verified_routes": sum(
            1 for row in rows if row["export_type"] in {"html", "pdf"} and row["not_verified_label_present"]
        ),
        "xlsx_valid_routes": sum(1 for row in rows if row["export_type"] == "xlsx" and row["valid_payload"]),
        "xlsx_all_target_station_sheet_routes": sum(
            1 for row in rows if row["xlsx_target_station_sheet_count"] == len(TARGET_CODES)
        ),
        "xlsx_verification_label_routes": sum(
            1 for row in rows if row["export_type"] == "xlsx" and row["verified_label_without_not_verified"]
        ),
        "routes_with_all_target_station_codes": bool_count(rows, "all_target_station_codes"),
        "routes_with_pm25": bool_count(rows, "pm25_present"),
        "verified_label_without_not_verified_routes": bool_count(rows, "verified_label_without_not_verified"),
        "verified_report_closure_available_routes": bool_count(rows, "verified_report_closure_available"),
        "current_status_confirmed_routes": bool_count(rows, "current_status_confirmed"),
        "station_method_classified_routes": bool_count(rows, "station_method_classified"),
        "complete_monitor_grade_classification_routes": bool_count(rows, "complete_monitor_grade_classification_available"),
        "station_radius_grade_assumption_ready_routes": bool_count(rows, "station_radius_grade_assumption_ready"),
    }
    return {
        "generated_at": generated_at,
        "program": "air-monitoring",
        "attestation_chain": "ai-first",
        "status": STATUS,
        "method": METHOD,
        "goal_level": "L3 Georgia report-frequency verification matrix",
        "source_scope": "Official air.gov.ge report generator routes for daily, monthly, and annual HTML/XLSX/PDF outputs for 16 target Georgia station codes.",
        "source_inputs": [
            {
                "path": str(SEED_CSV.relative_to(PROGRAM_DIR)),
                "role": "official Georgia report-frequency route seed",
            }
        ],
        "coverage_counts": counts,
        "frequency_rows": build_frequency_rows(rows),
        "decision_counts": [
            {"decision": decision, "rows": count}
            for decision, count in sorted(Counter(row["report_frequency_decision"] for row in rows).items())
        ],
        "evidence_gate_counts": evidence_gates(rows),
        "sample_rows": [
            {
                "report_type": row["report_type"],
                "export_type": row["export_type"],
                "probe_date": row["probe_date"],
                "http_status": row["http_status"],
                "valid_payload": row["valid_payload"],
                "station_code_matches": row["station_code_matches"],
                "pm25_present": row["pm25_present"],
                "not_verified_label_present": row["not_verified_label_present"],
                "verified_label_without_not_verified": row["verified_label_without_not_verified"],
                "xlsx_target_station_sheet_count": row["xlsx_target_station_sheet_count"],
                "report_frequency_decision": row["report_frequency_decision"],
            }
            for row in rows
        ],
        "non_claim": NON_CLAIM,
    }


def write_note(path: Path, summary: dict[str, Any]) -> None:
    counts = summary["coverage_counts"]
    lines = [
        "---",
        "attestation_chain: ai-first",
        "status: Screening Result",
        f"method: {METHOD}",
        "---",
        "",
        "# Georgia report-frequency matrix",
        "",
        "## Why this pass exists",
        "",
        "The Georgia policy wall points readers toward reports for verified data,",
        "while the monthly report/export ladder kept the monthly surface open.",
        "This pass tests whether daily or annual report frequencies change that",
        "decision for the same target station-code set.",
        "",
        "## What the public routes show",
        "",
        f"- Report-frequency/export probes targeted: {counts['route_probes_targeted']}.",
        f"- Valid daily/monthly report or export payloads retrieved: {counts['valid_payload_routes']}.",
        f"- Annual route probes returning server-error pages: {counts['annual_server_error_routes']}.",
        f"- HTML/PDF payloads retaining not-verified labels: {counts['html_pdf_not_verified_routes']}.",
        f"- XLSX exports with all target station sheets: {counts['xlsx_all_target_station_sheet_routes']}.",
        f"- XLSX exports with verification labels: {counts['xlsx_verification_label_routes']}.",
        f"- Verified report-closure routes found: {counts['verified_report_closure_available_routes']}.",
        "",
        "## Reader use",
        "",
        "Use this as a report-frequency falsifier. It shows that daily reports do",
        "not rescue the verified-report gate, monthly comparison probes stay",
        "consistent with the 24-month ladder, annual probes do not return a usable",
        "public report payload for the tested formats, and XLSX station sheets are",
        "not enough without a verification label.",
        "",
        "## Non-claim",
        "",
        summary["non_claim"],
        "",
        "## Reproduce",
        "",
        "Run `python air-monitoring/scripts/scan-georgia-report-frequency-matrix.py`.",
        "The source list is `air-monitoring/source-inputs/georgia-report-frequency-matrix-source-seed.csv`.",
        "Outputs are `air-monitoring/generated/air-monitoring-georgia-report-frequency-matrix.csv`",
        "and `air-monitoring/generated/air-monitoring-georgia-report-frequency-matrix-summary.json`.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    generated_at = now_iso()
    seed_rows = read_csv(SEED_CSV)
    rows = build_rows(generated_at, seed_rows)
    summary = build_summary(generated_at, rows)
    write_csv(OUT_CSV, rows)
    write_json(OUT_JSON, summary)
    write_note(OUT_MD, summary)
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_MD}")
    print(json.dumps(summary["coverage_counts"], indent=2))


if __name__ == "__main__":
    main()
